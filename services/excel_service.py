import uuid
import re
from datetime import datetime, timezone

from db.models import Product


def _clean(val):
    if val is None:
        return ""
    s = str(val).strip()
    return s


def _float(val):
    if val is None or str(val).strip() == "":
        return 0.0
    try:
        return float(val)
    except:
        return 0.0


def _guess_sku(name, group=""):
    import hashlib

    parts = re.sub(r"[^A-Za-z0-9/]", "_", name.strip().upper())[:30]
    grp = re.sub(r"[^A-Za-z0-9]", "", group.strip().upper())[:8]
    base = f"{grp}-{parts}" if grp else parts
    suffix = hashlib.md5(name.encode()).hexdigest()[:4].upper()
    return f"{base}_{suffix}"


def _get(row, idx):
    if idx is None:
        return None
    if 0 <= idx < len(row):
        return row[idx]
    return None


def _upsert_product(db, name, category, unit, group="", stock=0.0, price=0.0):
    if not name:
        return None, False
    sku = _guess_sku(name, group)
    existing = (
        db.query(Product)
        .filter(Product.name == name, Product.category == category)
        .first()
    )
    if existing:
        existing.current_stock = stock
        if price:
            existing.unit_price = price
        existing.unit = unit or existing.unit
        return existing, True
    product = Product(
        id=str(uuid.uuid4()),
        sku=sku,
        name=name,
        category=category or "Uncategorized",
        unit=unit or "Pcs",
        current_stock=stock,
        unit_price=price,
        created_at=datetime.now(timezone.utc),
    )
    db.add(product)
    return product, False


def _parse_stock_register(wb, db):
    created = 0
    updated = 0
    errors = []
    seen = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [_clean(h) for h in (rows[0] or [])]
        col_map = {h: i for i, h in enumerate(headers)}

        name_col = col_map.get("Item Name")
        cat_col = col_map.get("Category")
        grp_col = col_map.get("Item Group")
        unit_col = col_map.get("Unit")
        open_col = col_map.get("Opening Stock")
        total_col = col_map.get("Total Stock")
        close_col = col_map.get("Closing Stock")
        utilized_col = col_map.get("Sale/Utilized Qty")
        purchase_col = col_map.get("Purchase Qty")

        for row in rows[1:]:
            try:
                name = _clean(_get(row, name_col) or "")
                if not name or name.lower() == "item name":
                    continue
                key = name.lower().strip()
                if key in seen:
                    continue
                seen.add(key)

                cat = _clean(_get(row, cat_col) or "Uncategorized")
                grp = _clean(_get(row, grp_col) or "")
                unit = _clean(_get(row, unit_col) or "Pcs")
                opening = _float(_get(row, open_col))
                total = _float(_get(row, total_col)) or opening
                closing = _float(_get(row, close_col)) or total
                stock = max(opening, total, closing)

                prod, was_updated = _upsert_product(db, name, cat, unit, grp, stock)
                if prod:
                    if was_updated:
                        updated += 1
                    else:
                        created += 1
            except Exception as e:
                errors.append(f"{sheet_name}: {e}")

    db.commit()
    return {"created": created, "updated": updated, "errors": errors}


def _parse_consumption(wb, db):
    created = 0
    updated = 0
    errors = []
    seen = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [_clean(h) for h in (rows[0] or [])]
        col_map = {h: i for i, h in enumerate(headers)}

        name_col = col_map.get("Material Name")
        unit_col = col_map.get("Unit")
        issued_col = col_map.get("Issued Qty")
        person_col = col_map.get("Person/Helper")

        for row in rows[1:]:
            try:
                name = _clean(_get(row, name_col) or "")
                if not name:
                    continue
                key = name.lower().strip()
                if key in seen:
                    continue
                seen.add(key)

                unit = _clean(_get(row, unit_col) or "Pcs")
                issued = _float(_get(row, issued_col))
                person = _clean(_get(row, person_col) or sheet_name)
                cat = f"Consumable - {person}"

                prod, was_updated = _upsert_product(db, name, cat, unit, "", issued)
                if prod:
                    if was_updated:
                        updated += 1
                    else:
                        created += 1
            except Exception as e:
                errors.append(f"{sheet_name}: {e}")

    db.commit()
    return {"created": created, "updated": updated, "errors": errors}


def _parse_yusuf_leth(wb, db):
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "updated": 0, "errors": ["Empty sheet"]}

    header_row = None
    for i, row in enumerate(rows):
        vals = [_clean(v) for v in (row or [])]
        if "DES. OF GOODS" in vals or "DES" in vals:
            header_row = i
            break

    if header_row is None:
        return {"created": 0, "updated": 0, "errors": ["Could not find header row"]}

    headers = [_clean(h) for h in (rows[header_row] or [])]
    col_map = {h: i for i, h in enumerate(headers)}

    desc_col = col_map.get("DES. OF GOODS")
    out_qty_col = col_map.get("OUT QTY.")
    in_qty_col = col_map.get("IN QTY.")
    rate_col = col_map.get("RATE")
    bal_col = col_map.get("BALANCE")

    created = 0
    updated = 0
    errors = []
    seen = set()

    for row in rows[header_row + 1 :]:
        try:
            name = _clean(_get(row, desc_col) or "")
            if not name:
                continue
            key = name.lower().strip()
            if key in seen:
                continue
            seen.add(key)

            out_qty = _float(_get(row, out_qty_col))
            in_qty = _float(_get(row, in_qty_col))
            rate = _float(_get(row, rate_col))
            bal = _float(_get(row, bal_col)) or max(out_qty, in_qty)
            stock = max(bal, out_qty)

            prod, was_updated = _upsert_product(
                db, name, "Yusuf Leth Stock", "Pcs", "", stock, rate
            )
            if prod:
                if was_updated:
                    updated += 1
                else:
                    created += 1
        except Exception as e:
            errors.append(f"Row: {e}")

    db.commit()
    return {"created": created, "updated": updated, "errors": errors}


def _parse_standard(rows, header_row_idx, db):
    headers = [_clean(h) for h in (rows[header_row_idx] or [])]
    col_map = {h: i for i, h in enumerate(headers)}
    sku_col = col_map.get("SKU")
    name_col = col_map.get("Product Name")
    cat_col = col_map.get("Category")
    unit_col = col_map.get("Unit")
    open_col = col_map.get("Opening Stock")
    price_col = col_map.get("Unit Price")

    created = 0
    updated = 0
    errors = []
    for row_idx, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        try:
            sku = _clean(_get(row, sku_col) or "")
            name = _clean(_get(row, name_col) or "")
            cat = _clean(_get(row, cat_col) or "")
            unit = _clean(_get(row, unit_col) or "")
            stock = _float(_get(row, open_col))
            price = _float(_get(row, price_col))
            existing = db.query(Product).filter_by(sku=sku).first()
            if existing:
                existing.name = name
                existing.category = cat
                existing.unit = unit
                existing.current_stock = stock
                existing.unit_price = price
                updated += 1
            else:
                product = Product(
                    id=str(uuid.uuid4()),
                    sku=sku,
                    name=name,
                    category=cat,
                    unit=unit,
                    current_stock=stock,
                    unit_price=price,
                    created_at=datetime.now(timezone.utc),
                )
                db.add(product)
                created += 1
        except Exception as e:
            errors.append(f"Row {row_idx}: {e}")

    db.commit()
    return {"created": created, "updated": updated, "errors": errors}


def parse_products_excel(file_bytes: bytes, db) -> dict:
    from openpyxl import load_workbook
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
        tmp.write(file_bytes)
        tmp.flush()
        wb = load_workbook(tmp.name, data_only=True)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"created": 0, "updated": 0, "errors": ["Empty sheet"]}

        # Skip blank title rows to find the actual header row
        header_row_idx = 0
        for i, row in enumerate(rows):
            vals = [_clean(v) for v in (row or [])]
            non_empty = [v for v in vals if v]
            if len(non_empty) >= 3:
                header_row_idx = i
                break
        headers = [_clean(h) for h in (rows[header_row_idx] or [])]

        yusuf_keywords = {"DES. OF GOODS", "OUT QTY.", "IN QTY."}

        hs = set(headers)

        # Each format is keyed off its own anchor column(s) — "Opening Stock" and
        # "Category" appear in both the stock-register and standard layouts, so
        # matching on those alone misclassifies standard sheets as stock registers.
        if "Item Name" in hs:
            result = _parse_stock_register(wb, db)
        elif "Material Name" in hs:
            result = _parse_consumption(wb, db)
        elif hs & yusuf_keywords:
            result = _parse_yusuf_leth(wb, db)
        elif "SKU" in hs and "Product Name" in hs:
            result = _parse_standard(rows, header_row_idx, db)
        else:
            return {
                "created": 0,
                "updated": 0,
                "errors": [f"Unknown format. Headers: {headers[:10]}"],
            }

        wb.close()
        return result
